import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
from pathlib import Path
import openpyxl

class ExcelMergerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Files Merger")
        self.root.geometry("600x400")
        
        self.folder_path = tk.StringVar()
        self.sheet_name = tk.StringVar()
        self.header_row = tk.StringVar(value="1")  # По умолчанию заголовки в первой строке
        self.selected_columns = []
        self.available_columns = []
        
        # Создаем и размещаем элементы интерфейса
        self.create_widgets()
        
    def create_widgets(self):
        # Frame для выбора папки
        folder_frame = tk.Frame(self.root)
        folder_frame.pack(pady=10, padx=10, fill="x")
        
        tk.Label(folder_frame, text="Папка с Excel файлами:").pack(side="left")
        tk.Entry(folder_frame, textvariable=self.folder_path, width=50).pack(side="left", padx=5)
        tk.Button(folder_frame, text="Выбрать", command=self.select_folder).pack(side="left")
        
        # Frame для выбора листа
        sheet_frame = tk.Frame(self.root)
        sheet_frame.pack(pady=5, padx=10, fill="x")
        
        tk.Label(sheet_frame, text="Имя листа Excel:").pack(side="left")
        tk.Entry(sheet_frame, textvariable=self.sheet_name, width=20).pack(side="left", padx=5)
        tk.Button(sheet_frame, text="Проверить лист", command=self.check_sheet).pack(side="left")
        
        # Frame для указания строки заголовков
        header_frame = tk.Frame(self.root)
        header_frame.pack(pady=5, padx=10, fill="x")
        
        tk.Label(header_frame, text="Строка с заголовками (нумерация с 1):").pack(side="left")
        tk.Entry(header_frame, textvariable=self.header_row, width=5).pack(side="left", padx=5)
        
        # Frame для списка столбцов
        columns_frame = tk.Frame(self.root)
        columns_frame.pack(pady=10, padx=10, fill="both", expand=True)
        
        tk.Label(columns_frame, text="Доступные столбцы:").pack()
        
        self.columns_listbox = tk.Listbox(columns_frame, selectmode="multiple", height=10)
        self.columns_listbox.pack(fill="both", expand=True)
        
        # Frame для кнопок
        buttons_frame = tk.Frame(self.root)
        buttons_frame.pack(pady=10)
        
        tk.Button(buttons_frame, text="Загрузить столбцы", command=self.load_columns).pack(side="left", padx=5)
        tk.Button(buttons_frame, text="Объединить файлы", command=self.merge_files).pack(side="left", padx=5)
        
    def select_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.folder_path.set(folder)
            self.sheet_name.set("")  # Очищаем поле листа
            self.columns_listbox.delete(0, tk.END)
            self.available_columns = []
    
    def check_sheet(self):
        folder = self.folder_path.get()
        sheet_name = self.sheet_name.get().strip()
        
        if not folder:
            messagebox.showerror("Ошибка", "Выберите папку с Excel файлами!")
            return
            
        if not sheet_name:
            messagebox.showerror("Ошибка", "Введите имя листа Excel!")
            return
            
        try:
            # Получаем список Excel файлов
            excel_files = [f for f in os.listdir(folder) if f.endswith(('.xlsx', '.xls'))]
            if not excel_files:
                messagebox.showerror("Ошибка", "В папке нет Excel файлов!")
                return
                
            # Проверяем наличие листа в каждом файле
            missing_files = []
            for file in excel_files:
                file_path = os.path.join(folder, file)
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                if sheet_name not in workbook.sheetnames:
                    missing_files.append(file)
                workbook.close()
                
            if missing_files:
                messagebox.showwarning("Предупреждение", 
                    f"Лист '{sheet_name}' отсутствует в файлах: {', '.join(missing_files)}.\n"
                    "Эти файлы будут пропущены при объединении.")
            else:
                messagebox.showinfo("Успех", f"Лист '{sheet_name}' найден во всех файлах.")
                
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при проверке листа: {str(e)}")
    
    def load_columns(self):
        folder = self.folder_path.get()
        sheet_name = self.sheet_name.get().strip()
        
        if not folder:
            messagebox.showerror("Ошибка", "Выберите папку с Excel файлами!")
            return
            
        if not sheet_name:
            messagebox.showerror("Ошибка", "Введите имя листа Excel!")
            return
            
        try:
            # Проверяем корректность ввода строки заголовков
            try:
                header_row = int(self.header_row.get()) - 1  # Переводим в 0-based индекс для pandas
                if header_row < 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Ошибка", "Введите корректный номер строки заголовков (положительное число)!")
                return
                
            # Получаем первый Excel файл для определения столбцов
            excel_files = [f for f in os.listdir(folder) if f.endswith(('.xlsx', '.xls'))]
            if not excel_files:
                messagebox.showerror("Ошибка", "В папке нет Excel файлов!")
                return
                
            # Проверяем наличие листа в первом файле
            first_file_path = os.path.join(folder, excel_files[0])
            workbook = openpyxl.load_workbook(first_file_path, read_only=True)
            if sheet_name not in workbook.sheetnames:
                messagebox.showerror("Ошибка", f"Лист '{sheet_name}' отсутствует в файле {excel_files[0]}!")
                workbook.close()
                return
            workbook.close()
            
            # Читаем первый файл для получения заголовков
            first_file = pd.read_excel(first_file_path, sheet_name=sheet_name, header=header_row)
            self.available_columns = list(first_file.columns)
            
            # Очищаем и заполняем listbox
            self.columns_listbox.delete(0, tk.END)
            for col in self.available_columns:
                self.columns_listbox.insert(tk.END, col)
                
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при загрузке столбцов: {str(e)}")
    
    def merge_files(self):
        folder = self.folder_path.get()
        sheet_name = self.sheet_name.get().strip()
        
        if not folder:
            messagebox.showerror("Ошибка", "Выберите папку с Excel файлами!")
            return
            
        if not sheet_name:
            messagebox.showerror("Ошибка", "Введите имя листа Excel!")
            return
            
        # Получаем выбранные столбцы
        selected_indices = self.columns_listbox.curselection()
        if not selected_indices:
            messagebox.showerror("Ошибка", "Выберите хотя бы один столбец!")
            return
            
        self.selected_columns = [self.available_columns[i] for i in selected_indices]
        
        # Проверяем корректность ввода строки заголовков
        try:
            header_row = int(self.header_row.get()) - 1  # Переводим в 0-based индекс для pandas
            if header_row < 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Ошибка", "Введите корректный номер строки заголовков (положительное число)!")
            return
            
        try:
            # Получаем все Excel файлы
            excel_files = [f for f in os.listdir(folder) if f.endswith(('.xlsx', '.xls'))]
            if not excel_files:
                messagebox.showerror("Ошибка", "В папке нет Excel файлов!")
                return
                
            # Создаем пустой DataFrame для результата
            result_df = pd.DataFrame()
            
            # Объединяем файлы
            for file in excel_files:
                file_path = os.path.join(folder, file)
                
                # Проверяем наличие листа
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                if sheet_name not in workbook.sheetnames:
                    messagebox.showwarning("Предупреждение", 
                        f"Лист '{sheet_name}' отсутствует в файле {file}. Пропускаем файл.")
                    workbook.close()
                    continue
                workbook.close()
                
                # Читаем данные из указанного листа
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
                
                # Проверяем наличие выбранных столбцов
                missing_cols = [col for col in self.selected_columns if col not in df.columns]
                if missing_cols:
                    messagebox.showwarning("Предупреждение", 
                        f"В файле {file} отсутствуют столбцы: {', '.join(missing_cols)}")
                    continue
                    
                # Выбираем только нужные столбцы
                df = df[self.selected_columns]
                # Добавляем столбец "Источник" с именем файла
                df['Источник'] = file
                result_df = pd.concat([result_df, df], ignore_index=True)
            
            if result_df.empty:
                messagebox.showerror("Ошибка", "Нет данных для объединения! Проверьте файлы и лист.")
                return
                
            # Сохраняем результат
            output_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Сохранить объединенный файл"
            )
            
            if output_path:
                result_df.to_excel(output_path, index=False)
                messagebox.showinfo("Успех", f"Файлы успешно объединены и сохранены в {output_path}")
                
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при объединении файлов: {str(e)}")

def main():
    root = tk.Tk()
    app = ExcelMergerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()